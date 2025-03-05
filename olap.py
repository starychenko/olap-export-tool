import os
import sys
import clr
import re
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import time
import datetime
from colorama import init, Fore, Back, Style
import threading
import itertools

# Імпорт для COM-інтерфейсу ADO (OLE DB)
try:
    import win32com.client
    HAS_PYWIN32 = True
except ImportError:
    HAS_PYWIN32 = False

# Ініціалізуємо colorama для кольорового виводу в консоль
init(autoreset=True)

# Завантажуємо змінні середовища з .env файлу
load_dotenv()

# Глобальні змінні для керування анімацією
animation_running = False
avg_query_time = None  # Середній час виконання запиту (ініціалізується при першому вимірі)

# Константи для методів автентифікації
AUTH_SSPI = "SSPI"
AUTH_LOGIN = "LOGIN"

# Додаємо шлях до Microsoft.AnalysisServices.AdomdClient.dll з .env
adomd_dll_path = os.getenv('ADOMD_DLL_PATH')
sys.path.append(adomd_dll_path)
clr.AddReference('Microsoft.AnalysisServices.AdomdClient')

from pyadomd import Pyadomd
import pandas as pd

# Функція для отримання імені поточного користувача Windows
def get_current_windows_user():
    """
    Повертає ім'я поточного користувача Windows.
    
    Функція використовує декілька методів для надійного визначення реального користувача,
    від імені якого виконується Windows-автентифікація (SSPI).
    
    Returns:
        str: Ім'я поточного користувача Windows
        
    Notes:
        1. Спочатку використовується os.getlogin(), який визначає користувача за поточною сесією.
        2. Якщо цей метод не вдається (наприклад, при запуску в деяких контейнерах або через 
           планувальник завдань), використовується системна змінна середовища USERNAME.
        3. У випадку, якщо і USERNAME не визначено, повертається 'Невідомий користувач'.
    """
    try:
        current_user = os.getlogin()
    except Exception:
        # Запасний варіант, якщо getlogin() не спрацює
        current_user = os.getenv('USERNAME', 'Невідомий користувач')
    return current_user

# Визначаємо поточний рік та тиждень для значень за замовчуванням
CURRENT_YEAR = datetime.datetime.now().year
CURRENT_WEEK = datetime.datetime.now().isocalendar()[1]  # Поточний номер тижня

# Функція для виводу часу
def get_current_time():
    return datetime.datetime.now().strftime('%H:%M:%S')

# Функція для виводу заголовків
def print_header(text):
    print(f"\n{Fore.CYAN}{Style.BRIGHT}{'=' * 80}")
    print(f"{Fore.CYAN}{Style.BRIGHT}== {text}")
    print(f"{Fore.CYAN}{Style.BRIGHT}{'=' * 80}")
    print() # Додаємо порожній рядок для кращої читабельності

# Функція для виводу інформаційних повідомлень з деталями
def print_info_detail(text, details=None):
    """
    Виводить інформаційне повідомлення з додатковими деталями
    
    Args:
        text (str): Основне повідомлення
        details (dict, optional): Словник з деталями у форматі ключ-значення
    """
    print(f"{Fore.GREEN}[{get_current_time()}] ℹ️ {text}")
    
    if details:
        for key, value in details.items():
            # Якщо значення є паролем, приховуємо його
            if 'password' in key.lower() or 'пароль' in key.lower():
                value = '********'
            print(f"   {Fore.CYAN}{key}: {Fore.WHITE}{value}")

# Функція для виводу детальної технічної помилки
def print_tech_error(text, error_obj=None):
    """
    Виводить технічну помилку з детальною інформацією
    
    Args:
        text (str): Основне повідомлення про помилку
        error_obj (Exception, optional): Об'єкт виключення для виводу деталей
    """
    print(f"{Fore.RED}[{get_current_time()}] 🛑 {text}")
    
    if error_obj:
        error_type = type(error_obj).__name__
        error_message = str(error_obj)
        
        print(f"   {Fore.RED}Тип помилки: {Fore.WHITE}{error_type}")
        print(f"   {Fore.RED}Повідомлення: {Fore.WHITE}{error_message}")
        
        # Якщо є traceback, виводимо останні 3 рядки стеку викликів
        if hasattr(error_obj, '__traceback__') and error_obj.__traceback__:
            import traceback
            tb_lines = traceback.format_tb(error_obj.__traceback__)
            if len(tb_lines) > 3:
                tb_lines = tb_lines[-3:]  # Останні 3 рядки
            
            print(f"   {Fore.RED}Стек викликів:")
            for line in tb_lines:
                print(f"   {Fore.YELLOW}{line.strip()}")

# Функція для виводу інформаційних повідомлень
def print_info(text):
    print(f"{Fore.GREEN}[{get_current_time()}] ℹ️ {text}")

# Функція для виводу попереджень
def print_warning(text):
    print(f"{Fore.YELLOW}[{get_current_time()}] ⚠️ {text}")

# Функція для виводу помилок
def print_error(text):
    print(f"{Fore.RED}[{get_current_time()}] ❌ {text}")

# Функція для виводу успішних операцій
def print_success(text):
    print(f"{Fore.GREEN}[{get_current_time()}] ✅ {text}")

# Функція для виводу прогресу
def print_progress(text):
    print(f"{Fore.BLUE}[{get_current_time()}] 🔄 {text}")

# Функція для форматування часу у вигляді години:хвилини:секунди
def format_time(seconds):
    """Форматує час у секундах до читабельного формату (години, хвилини, секунди)"""
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    
    # Показуємо тільки ті одиниці виміру, які більше 0
    if hours > 0:
        return f"{int(hours)} год {int(minutes)} хв {seconds:.2f} сек"
    elif minutes > 0:
        return f"{int(minutes)} хв {seconds:.2f} сек"
    else:
        return f"{seconds:.2f} сек"

# Клас для відстеження прогресу та часу виконання завдання
class TimeTracker:
    """Клас для відстеження часу виконання та прогнозування завершення"""
    def __init__(self, total_items):
        """Ініціалізує трекер часу"""
        self.total_items = total_items
        self.processed_items = 0
        self.start_time = time.time()
        self.elapsed_times = []  # Час на обробку кожного елемента (без пауз)
        self.waiting_times = []  # Час пауз між елементами
        self.last_item_end_time = self.start_time  # Час завершення обробки останнього елемента
        self.currently_waiting = False  # Флаг, що показує, чи в режимі очікування ми зараз
    
    def start_waiting(self):
        """Позначає початок періоду очікування"""
        self.currently_waiting = True
        self.wait_start_time = time.time()
    
    def end_waiting(self):
        """Позначає кінець періоду очікування і зберігає час очікування"""
        if self.currently_waiting:
            wait_time = time.time() - self.wait_start_time
            self.waiting_times.append(wait_time)
            self.currently_waiting = False
    
    def update(self, items_processed=1):
        """Оновлює статус обробки після завершення елемента"""
        current_time = time.time()
        
        # Якщо ми були в режимі очікування, завершуємо його
        if self.currently_waiting:
            self.end_waiting()
        
        # Розраховуємо час на останній елемент (без пауз)
        if self.processed_items == 0:
            # Для першого елемента просто від початку до поточного часу
            processing_time = current_time - self.start_time
        else:
            # Для наступних елементів від кінця останнього елемента
            processing_time = current_time - self.last_item_end_time
            # Віднімаємо час очікування, якщо такий був
            if self.waiting_times:
                processing_time -= self.waiting_times[-1]
        
        # Зберігаємо час обробки і оновлюємо час закінчення останнього елемента
        self.elapsed_times.append(processing_time)
        self.last_item_end_time = current_time
        self.processed_items += items_processed
    
    def get_elapsed_time(self):
        """Повертає час, що минув з початку виконання"""
        return time.time() - self.start_time
    
    def get_processing_time(self):
        """Повертає час, витрачений на обробку даних (без пауз)"""
        return sum(self.elapsed_times) if self.elapsed_times else 0
    
    def get_waiting_time(self):
        """Повертає час, витрачений на паузи між запитами"""
        return sum(self.waiting_times) if self.waiting_times else 0
    
    def get_remaining_processing_time(self):
        """Прогнозує час обробки даних, що залишився (без пауз)"""
        if not self.elapsed_times or self.processed_items == 0:
            return None  # Не можемо спрогнозувати без даних
        
        # Використовуємо останні 5 елементів (або всі наявні, якщо їх менше) для більш точного прогнозу
        num_items_to_use = min(5, len(self.elapsed_times))
        recent_times = self.elapsed_times[-num_items_to_use:]
        
        # Простий розрахунок середнього часу на елемент
        avg_time_per_item = sum(recent_times) / len(recent_times)
        
        # Діагностичний вивід перенесений у get_progress_info()
        
        # Якщо оброблено мало елементів, додаємо коефіцієнт безпеки
        if len(self.elapsed_times) < 5 or self.processed_items < self.total_items * 0.1:
            # Додаємо коефіцієнт, який залежить від кількості оброблених елементів
            if len(self.elapsed_times) == 1:
                safety_factor = 1.2  # +20% для першого елемента
            elif len(self.elapsed_times) < 3:
                safety_factor = 1.1  # +10% для 2-3 елементів
            else:
                safety_factor = 1.05  # +5% для 4-5 елементів
                
            avg_time_per_item *= safety_factor
        
        # Кількість елементів, що залишилося обробити
        remaining_items = self.total_items - self.processed_items
        
        # Прогноз часу, що залишився на обробку
        remaining_time = avg_time_per_item * remaining_items
        
        return remaining_time
    
    def get_remaining_wait_time(self):
        """Прогнозує час очікування, що залишився"""
        # Отримуємо час очікування з конфігурації замість жорстко закодованого значення
        wait_time_per_item = int(os.getenv('QUERY_TIMEOUT', 30))  # Значення з .env або за замовчуванням 30 сек
        
        # Кількість елементів, що залишилося обробити (мінус 1, бо після останнього елемента немає очікування)
        remaining_items = max(0, self.total_items - self.processed_items - 1)
        
        return wait_time_per_item * remaining_items
    
    def get_remaining_time(self):
        """Прогнозує загальний час, що залишився до завершення (обробка + очікування)"""
        processing_time = self.get_remaining_processing_time()
        if processing_time is None:
            return None
            
        waiting_time = self.get_remaining_wait_time()
        return processing_time + waiting_time
    
    def get_percentage_complete(self):
        """Повертає відсоток виконання завдання"""
        return (self.processed_items / self.total_items) * 100 if self.total_items > 0 else 0
    
    def get_total_time(self):
        """Прогнозує загальний час на виконання"""
        remaining = self.get_remaining_time()
        if remaining is None:
            return self.get_elapsed_time()  # Повертаємо лише час, що пройшов
        return self.get_elapsed_time() + remaining
        
    def get_progress_info(self):
        """Повертає інформацію про прогрес у зручному форматі"""
        # Отримуємо базові значення
        elapsed = self.get_elapsed_time()
        processing_time = self.get_processing_time()
        waiting_time = self.get_waiting_time()
        remaining_processing = self.get_remaining_processing_time()
        remaining_waiting = self.get_remaining_wait_time()
        remaining_total = self.get_remaining_time()
        total = self.get_total_time()
        percentage = self.get_percentage_complete()
        
        # Розрахунок діагностичних значень для виводу
        debug_output = os.getenv('DEBUG', 'false').lower() in ('true', '1', 'yes')  # Контроль через змінну середовища
        if debug_output and self.elapsed_times and self.processed_items > 0:
            # Використовуємо останні 5 елементів для аналізу
            num_items_to_use = min(5, len(self.elapsed_times))
            recent_times = self.elapsed_times[-num_items_to_use:]
            
            # Середній час на елемент (тільки обробка)
            avg_processing_time = sum(recent_times) / len(recent_times)
            print(f"DEBUG: Середній час на обробку елемента: {avg_processing_time:.2f} сек", file=sys.stderr)
            print(f"DEBUG: Останні {len(recent_times)} виміри часу: {[round(t, 2) for t in recent_times]}", file=sys.stderr)
            
            # Якщо є дані про час очікування
            if self.waiting_times:
                avg_waiting_time = sum(self.waiting_times) / len(self.waiting_times)
                print(f"DEBUG: Середній час очікування: {avg_waiting_time:.2f} сек", file=sys.stderr)
            
            # Застосування коефіцієнта безпеки
            safety_factor = 1.0
            if len(self.elapsed_times) < 5 or self.processed_items < self.total_items * 0.1:
                if len(self.elapsed_times) == 1:
                    safety_factor = 1.2
                elif len(self.elapsed_times) < 3:
                    safety_factor = 1.1
                else:
                    safety_factor = 1.05
                
                print(f"DEBUG: Застосовано коефіцієнт безпеки {safety_factor:.2f}x до часу обробки", file=sys.stderr)
            
            # Інформація про елементи, що залишились
            remaining_items = self.total_items - self.processed_items
            print(f"DEBUG: Залишилось елементів: {remaining_items}", file=sys.stderr)
            
            # Деталізація прогнозів з форматованим часом
            if remaining_processing is not None:
                print(f"DEBUG: Прогноз часу обробки: {format_time(remaining_processing)} ({remaining_processing:.2f} сек)", file=sys.stderr)
                print(f"DEBUG: Прогноз часу очікування: {format_time(remaining_waiting)} ({remaining_waiting:.2f} сек)", file=sys.stderr)
                print(f"DEBUG: Загальний прогноз часу: {format_time(remaining_total)} ({remaining_total:.2f} сек)", file=sys.stderr)
        
        # Формуємо рядок виводу для користувача
        info = f"Прогрес: {percentage:.1f}% ({self.processed_items}/{self.total_items})\n"
        info += f"Минуло: {format_time(elapsed)}"
        
        if remaining_total is not None:
            # Додаємо примітку щодо точності прогнозу для перших елементів
            accuracy_note = ""
            if len(self.elapsed_times) == 1:
                accuracy_note = " (дуже приблизно)"
            elif len(self.elapsed_times) < 3:
                accuracy_note = " (орієнтовно)"
            
            info += f" | Залишилось: {format_time(remaining_total)}{accuracy_note}"
            info += f" | Всього: {format_time(total)}{accuracy_note}"
            
            # Додаємо додаткову діагностичну інформацію, якщо потрібно 
            if debug_output and processing_time > 0 and waiting_time > 0:
                processing_percentage = (processing_time / (processing_time + waiting_time)) * 100
                total_processing_time = processing_time + remaining_processing if remaining_processing is not None else processing_time
                total_waiting_time = waiting_time + remaining_waiting if remaining_waiting is not None else waiting_time
                info += f"\nDEBUG: Час обробки: {format_time(total_processing_time)} ({processing_percentage:.1f}%) | Час очікування: {format_time(total_waiting_time)} ({100-processing_percentage:.1f}%)"
        
        return info

# Функція для анімованого індикатора завантаження
def loading_spinner(description, estimated_time=None):
    """Функція для відображення анімованого індикатора завантаження"""
    global animation_running
    animation_running = True
    
    # Символи для анімації
    spinner = itertools.cycle(['⣾', '⣽', '⣻', '⢿', '⡿', '⣟', '⣯', '⣷'])
    
    # Початковий час для відображення тривалості
    start_time = time.time()
    
    # Відображаємо анімацію поки вона активна
    while animation_running:
        elapsed = time.time() - start_time
        # Використовуємо нашу функцію format_time для форматування часу
        elapsed_str = format_time(elapsed)
        
        # Базовий рядок з інформацією та часом виконання поточного запиту
        message = f"{Fore.BLUE}[{get_current_time()}] {next(spinner)} {description} | Час: {elapsed_str}"
        
        sys.stdout.write(f"\r{message}")
        sys.stdout.flush()
        time.sleep(0.1)
    
    # Очищаємо останній рядок анімації (використовуємо довжину останнього повідомлення)
    sys.stdout.write("\r" + " " * len(message) + "\r")
    sys.stdout.flush()
    # Додаємо новий рядок для відокремлення від наступного повідомлення
    print()

# Функція для генерації переліку тижнів за періодом
def generate_year_week_pairs(start_period, end_period, available_weeks):
    """Генерує список пар (рік, тиждень) в заданому діапазоні, враховуючи доступні тижні у кубі"""
    # Парсимо початковий і кінцевий періоди (формат РРРР-ТТ)
    try:
        start_year, start_week = map(int, start_period.split('-'))
        end_year, end_week = map(int, end_period.split('-'))
    except (ValueError, AttributeError):
        print_error(f"Невірний формат періодів. Використовуйте формат РРРР-ТТ")
        return []
    
    # Перевіряємо коректність введених даних
    current_year = datetime.datetime.now().year
    min_year = current_year - 3
    max_year = current_year

    if start_year < min_year or end_year > max_year:
        print_error(f"Невірні значення року (має бути між {min_year} та {max_year})")
        return []
    
    if start_year > end_year or (start_year == end_year and start_week > end_week):
        print_error(f"Початковий період має бути раніше за кінцевий")
        return []
    
    # Створюємо словник доступних тижнів для швидкого пошуку
    available_dict = {(year, week): True for year, week in available_weeks}
    
    # Фільтруємо за доступними тижнями
    filtered_pairs = []
    
    # Генеруємо всі потенційні пари
    all_pairs = []
    current_year = start_year
    current_week = start_week
    
    while current_year < end_year or (current_year == end_year and current_week <= end_week):
        all_pairs.append((current_year, current_week))
        current_week += 1
        # Якщо перейшли до наступного року
        if current_week > 53:  # Використовуємо 53 як максимальне значення тижня
            current_week = 0   # Починаємо з тижня 0, якщо він існує
            current_year += 1
    
    # Фільтруємо пари за наявністю в кубі
    for year, week in all_pairs:
        if (year, week) in available_dict:
            filtered_pairs.append((year, week))
    
    if len(filtered_pairs) == 0:
        print_warning(f"Не знайдено доступних тижнів у вказаному діапазоні")
    else:
        print_info(f"Знайдено {len(filtered_pairs)} тижнів у вказаному діапазоні")
    
    return filtered_pairs

# Функція для отримання рядка підключення до OLAP
def get_connection_string():
    """
    Повертає рядок підключення до OLAP сервера на основі налаштувань з .env
    
    Returns:
        tuple: (connection_string, auth_details)
            - connection_string (str): Рядок підключення до OLAP серверу
            - auth_details (dict): Словник з деталями автентифікації
    
    Notes:
        При Windows-автентифікації (SSPI) ім'я поточного користувача визначається 
        за допомогою функції get_current_windows_user(), яка використовує 
        os.getlogin() та запасний варіант os.getenv('USERNAME').
    """
    # Читаємо базові параметри
    server = os.getenv('OLAP_SERVER')
    database = os.getenv('OLAP_DATABASE')
    
    # Читаємо метод автентифікації з .env
    auth_method = os.getenv('OLAP_AUTH_METHOD', AUTH_SSPI).upper()  # За замовчуванням SSPI
    
    # Формуємо базову частину рядка підключення
    connection_string = f"Provider=MSOLAP;Data Source={server};Initial Catalog={database};"
    
    # Додаємо параметри автентифікації
    if auth_method == AUTH_SSPI:
        # Windows-автентифікація
        connection_string += "Integrated Security=SSPI;"
        auth_details = {
            "Метод автентифікації": "Windows-автентифікація (SSPI)",
            "Поточний користувач": get_current_windows_user()
        }
    elif auth_method == AUTH_LOGIN:
        # Автентифікація за логіном/паролем
        user = os.getenv('OLAP_USER')
        password = os.getenv('OLAP_PASSWORD')
        
        if not user or not password:
            print_warning("Обрано автентифікацію за логіном/паролем, але дані не вказані. Використовуємо SSPI.")
            connection_string += "Integrated Security=SSPI;"
            auth_details = {
                "Метод автентифікації": "Windows-автентифікація (SSPI) - автоматично",
                "Поточний користувач": get_current_windows_user(),
                "Причина": "Логін або пароль не вказані"
            }
        else:
            connection_string += f"User ID={user};Password={password};Persist Security Info=True;Update Isolation Level=2;"
            auth_details = {
                "Метод автентифікації": "Логін/пароль",
                "Користувач": user,
                "Пароль": password  # Буде приховано у виводі
            }
    else:
        # Невідомий метод автентифікації, використовуємо SSPI
        print_warning(f"Невідомий метод автентифікації '{auth_method}'. Використовуємо SSPI.")
        connection_string += "Integrated Security=SSPI;"
        auth_details = {
            "Метод автентифікації": "Windows-автентифікація (SSPI) - автоматично",
            "Поточний користувач": get_current_windows_user(),
            "Причина": f"Невідомий метод автентифікації: {auth_method}"
        }
    
    return connection_string, auth_details

# Функція для підключення через ADO (OLE DB) з використанням pywin32
def connect_using_ado(connection_string, auth_details):
    """
    Підключається до OLAP сервера через ADO (OLE DB) за допомогою pywin32
    
    Args:
        connection_string (str): Рядок підключення до OLAP серверу
        auth_details (dict): Словник з деталями автентифікації
        
    Returns:
        tuple: (connection, cursor) - ADO з'єднання та курсор для запитів
        
    Notes:
        Використовує pywin32 для створення COM-об'єкта ADODB.Connection.
        Цей метод дозволяє надійно використовувати автентифікацію за логіном/паролем.
    """
    if not HAS_PYWIN32:
        print_error("Бібліотека pywin32 не знайдена. Встановіть її командою: pip install pywin32")
        return None, None
    
    try:
        print_info_detail(f"Підключення до OLAP сервера {os.getenv('OLAP_SERVER')} через ADO...", auth_details)
        
        # Створюємо COM-об'єкт для ADO підключення
        connection = win32com.client.Dispatch(r'ADODB.Connection')
        connection.Open(connection_string)
        
        # Створюємо COM-об'єкт для команд
        command = win32com.client.Dispatch(r'ADODB.Command')
        command.ActiveConnection = connection
        
        # Створюємо обгортку-курсор для сумісності з іншим кодом
        cursor = AdoCursor(command)
        
        print_success(f"Підключення до OLAP сервера через ADO успішно встановлено")
        return connection, cursor
    except Exception as e:
        print_tech_error(f"Помилка підключення до OLAP сервера через ADO", e)
        
        # Додаткова інформація про можливі причини помилки
        if "Login failed" in str(e) or "логін" in str(e).lower():
            print_warning("Можлива причина: Неправильний логін або пароль")
            print_info("Рекомендація: Перевірте значення OLAP_USER та OLAP_PASSWORD у файлі .env")
        elif "provider" in str(e).lower():
            print_warning("Можлива причина: Проблеми з провайдером MSOLAP")
            print_info("Рекомендації:")
            print(f"   {Fore.CYAN}1. Перевірте наявність встановленого SQL Server або Analysis Services")
            print(f"   {Fore.CYAN}2. Перевірте версію провайдера MSOLAP")
        
        return None, None

# Клас-обгортка для забезпечення сумісності ADO з іншим кодом
class AdoCursor:
    """
    Клас-обгортка для ADO команди, щоб забезпечити спільний інтерфейс з pyadomd
    """
    def __init__(self, command):
        self.command = command
        self.rows = None
        self.columns = None
    
    def execute(self, query):
        """Виконує MDX запит"""
        self.command.CommandText = query
        self.command.CommandType = 1  # adCmdText
        self.recordset = self.command.Execute()[0]
    
    def fetchall(self):
        """Отримує всі результати запиту"""
        if not self.recordset:
            return []
        
        # Отримуємо поля (стовпці)
        fields = {}
        for i in range(self.recordset.Fields.Count):
            field = self.recordset.Fields(i)
            fields[i] = field.Name
        
        self.columns = list(fields.values())
        
        # Отримуємо всі рядки
        rows = []
        if not self.recordset.EOF:
            self.recordset.MoveFirst()
            while not self.recordset.EOF:
                row = []
                for i in range(self.recordset.Fields.Count):
                    row.append(self.recordset.Fields(i).Value)
                rows.append(row)
                self.recordset.MoveNext()
        
        self.rows = rows
        return rows
    
    def fetchone(self):
        """Отримує один рядок результатів"""
        if not self.recordset or self.recordset.EOF:
            return None
        
        row = []
        for i in range(self.recordset.Fields.Count):
            row.append(self.recordset.Fields(i).Value)
        
        self.recordset.MoveNext()
        return row
    
    def get_column_names(self):
        """Повертає імена стовпців"""
        if not self.columns:
            return []
        return self.columns

# Функція для підключення до OLAP сервера
def connect_to_olap(connection_string=None, auth_details=None):
    """Підключається до OLAP сервера і повертає з'єднання"""
    if connection_string is None:
        connection_string, auth_details = get_connection_string()
    
    # Визначаємо метод автентифікації
    auth_method = os.getenv('OLAP_AUTH_METHOD', AUTH_SSPI).upper()
    
    try:
        # Якщо використовується LOGIN автентифікація - використовуємо ADO через pywin32
        # Якщо використовується SSPI автентифікація - використовуємо ADOMD.NET
        if auth_method == AUTH_LOGIN and os.getenv('OLAP_USER') and os.getenv('OLAP_PASSWORD'):
            # Перевіряємо, чи встановлено pywin32
            if not HAS_PYWIN32:
                print_warning("Обрано автентифікацію за логіном/паролем (LOGIN), але бібліотека pywin32 не встановлена.")
                print_info("Рекомендація: Встановіть pywin32 командою: pip install pywin32")
                print_warning("Буде використано ADOMD.NET, але автентифікація за логіном/паролем може не спрацювати")
            else:
                # Використовуємо ADO через pywin32
                print_info(f"Використовуємо підключення через ADO (OLE DB) для автентифікації за логіном/паролем")
                ado_connection, cursor = connect_using_ado(connection_string, auth_details)
                
                if ado_connection:
                    # Створюємо обгортку для сумісності з іншими функціями
                    connection_wrapper = type('ADOConnectionWrapper', (), {
                        'cursor': lambda self: cursor,
                        'close': lambda self: ado_connection.Close(),
                        '_ado_connection': ado_connection  # Зберігаємо посилання на оригінальне підключення
                    })
                    return connection_wrapper()
                
                # Якщо ADO підключення не вдалося, повідомляємо про помилку
                print_error("Не вдалося встановити ADO підключення. Перевірте параметри підключення.")
                print_warning("Спробуємо використати ADOMD.NET, але автентифікація за логіном/паролем може не спрацювати.")
        
        # В інших випадках використовуємо ADOMD.NET (працює добре для Windows-автентифікації)
        print_info_detail(f"Підключення до OLAP сервера {os.getenv('OLAP_SERVER')} через ADOMD.NET...", auth_details)
        
        # Інформація про версію провайдера та шлях до DLL
        print_info(f"Шлях до ADOMD.NET: {adomd_dll_path}")
        dll_exists = os.path.exists(adomd_dll_path)
        if not dll_exists:
            print_warning("Шлях до ADOMD.NET не знайдено! Перевірте налаштування ADOMD_DLL_PATH у файлі .env")
        else:
            dll_files = [f for f in os.listdir(adomd_dll_path) if f.lower().endswith('.dll')]
            adomd_files = [f for f in dll_files if 'adomd' in f.lower()]
            if adomd_files:
                print_info(f"Знайдено ADOMD.NET файли: {', '.join(adomd_files)}")
            else:
                print_warning("У вказаному каталозі не знайдено файлів ADOMD.NET!")
                
        connection = Pyadomd(connection_string)
        connection.open()
        
        print_success(f"Підключення до OLAP сервера успішно встановлено")
        return connection
    except Exception as e:
        print_tech_error(f"Помилка підключення до OLAP сервера", e)
        
        # Додаткова інформація про можливі причини помилки
        if "Login failed" in str(e) or "логін" in str(e).lower():
            print_warning("Можлива причина: Неправильний логін або пароль")
            print_info("Рекомендація: Перевірте значення OLAP_USER та OLAP_PASSWORD у файлі .env")
        elif "provider" in str(e).lower():
            print_warning("Можлива причина: Проблеми з провайдером")
            print_info("Рекомендації:")
            if "ADOMD" in str(e):
                print(f"   {Fore.CYAN}1. Перевірте шлях до ADOMD.NET у змінній ADOMD_DLL_PATH у файлі .env")
                print(f"   {Fore.CYAN}2. Встановіть або перевстановіть Microsoft SQL Server Management Studio")
            else:
                print(f"   {Fore.CYAN}1. Перевірте наявність встановленого SQL Server або Analysis Services")
                print(f"   {Fore.CYAN}2. Спробуйте інший метод автентифікації")
        elif "Data Source" in str(e) or "сервер" in str(e).lower():
            print_warning("Можлива причина: Неправильна адреса сервера або сервер недоступний")
            print_info("Рекомендації:")
            print(f"   {Fore.CYAN}1. Перевірте значення OLAP_SERVER у файлі .env")
            print(f"   {Fore.CYAN}2. Перевірте, чи доступний сервер {os.getenv('OLAP_SERVER')} з вашої мережі")
            print(f"   {Fore.CYAN}3. Спробуйте виконати ping {os.getenv('OLAP_SERVER')}")
        elif "SSPI" in str(e):
            print_warning("Можлива причина: Проблеми з Windows-автентифікацією")
            print_info("Рекомендації:")
            if not HAS_PYWIN32:
                print(f"   {Fore.CYAN}1. Встановіть pywin32 для використання автентифікації за логіном/паролем: pip install pywin32")
            print(f"   {Fore.CYAN}2. Змініть метод автентифікації на LOGIN та вкажіть логін і пароль у файлі .env")
            print(f"   {Fore.CYAN}3. Перевірте, чи має ваш користувач {get_current_windows_user()} доступ до OLAP-кубу")
            
        # Вивід технічних деталей для відладки
        print_info("Технічні деталі для відладки:")
        print(f"   {Fore.CYAN}Рядок підключення: {Fore.WHITE}{connection_string.replace(os.getenv('OLAP_PASSWORD', ''), '********') if os.getenv('OLAP_PASSWORD') else connection_string}")
        
        return None

# Функція для виконання MDX-запиту і отримання результатів
def run_mdx_query(connection, reporting_period):
    """Виконує MDX-запит для заданого періоду і повертає результати"""
    # Парсимо період (формат РРРР-ТТ)
    try:
        year_num, week_num = map(int, reporting_period.split('-'))
    except (ValueError, AttributeError):
        print_error(f"Невірний формат періоду: {reporting_period}. Використовуйте формат РРРР-ТТ")
        return []
    
    # Отримуємо фільтр для запиту
    filter_fg1_name = os.getenv('FILTER_FG1_NAME')
    
    # Формуємо шлях для збереження результатів
    result_dir = "result"
    year_dir = os.path.join(result_dir, str(year_num))
    
    # Перевіряємо і створюємо папку для року, якщо вона не існує
    if not os.path.exists(year_dir):
        os.makedirs(year_dir)
        print_info(f"Створено директорію '{year_dir}'")
    
    # Формуємо назву файлу з ведучим нулем для тижня
    filename = f"{year_num}-{week_num:02d}.xlsx"
    # Повний шлях до файлу
    filepath = os.path.join(year_dir, filename)
    
    # Виводимо інформацію про запит
    print_info(f"Формування MDX запиту з параметрами:")
    print(f"   {Fore.CYAN}Рік:      {Fore.WHITE}{year_num}")
    print(f"   {Fore.CYAN}Тиждень:  {Fore.WHITE}{week_num}")
    print(f"   {Fore.CYAN}Фільтр:   {Fore.WHITE}{filter_fg1_name}")
    
    # Формуємо запит із використанням змінних для року та тижня
    query = f"""
    /* START QUERY BUILDER */
    EVALUATE
    SUMMARIZECOLUMNS(
        'Calendar'[calendar_date],
        Goods[fg1_name],
        Goods[fg2_name],
        Goods[fg3_name],
        Goods[fg4_name],
        Goods[articul],
        Goods[articul_name],
        Goods[producer_name],
        Agents_hybrid[name],
        Markets[doc_prefix_original],
        Channel_type[sell_channel_type_name],
        Price_types[name],
        Price_types[is_tender],
        Doc_types[name],
        Credit_products[payment_code],
        Credit_products[payment_typ],
        Credit_products[product_types],
        Credit_products[bank_name],
        Credit_products[bank_credit_product_code],
        Credit_products[product_name],
        Credit_products[payment_count],
        Promo[promo_type_name],
        Promo[basis],
        KEEPFILTERS( TREATAS( {{{year_num}}}, 'Calendar'[year_num] )),
        KEEPFILTERS( TREATAS( {{{week_num}}}, 'Calendar'[week_num] )),
        KEEPFILTERS( TREATAS( {{"{filter_fg1_name}"}}, Goods[fg1_name] )),
        "Реалізація, к-сть", [sell_qty],
        "Реалізація, грн.", [sell_amount_nds],
        "Реалізація ЦЗ, грн.", [buy_amount_nds],
        "Дохід, грн.", [profit_amount_nds],
        "Отримані бонуси", [bonus_obtained_amount],
        "Використані бонуси", [bonus_used_amount],
        "Комісія по кредитам", [credit_commission_amount]
    )
    ORDER BY 
        'Calendar'[calendar_date] ASC,
        Goods[fg1_name] ASC,
        Goods[fg2_name] ASC,
        Goods[fg3_name] ASC,
        Goods[fg4_name] ASC,
        Goods[articul] ASC,
        Goods[articul_name] ASC,
        Goods[producer_name] ASC,
        Agents_hybrid[name] ASC,
        Markets[doc_prefix_original] ASC,
        Channel_type[sell_channel_type_name] ASC,
        Price_types[name] ASC,
        Price_types[is_tender] ASC,
        Doc_types[name] ASC,
        Credit_products[payment_code] ASC,
        Credit_products[payment_typ] ASC,
        Credit_products[product_types] ASC,
        Credit_products[bank_name] ASC,
        Credit_products[bank_credit_product_code] ASC,
        Credit_products[product_name] ASC,
        Credit_products[payment_count] ASC,
        Promo[promo_type_name] ASC,
        Promo[basis] ASC
    /* END QUERY BUILDER */
    """
    
    print_progress(f"Виконання запиту до OLAP-кубу...")
    query_start_time = time.time()
    global animation_running
    
    try:
        cursor = connection.cursor()
        
        # Виконуємо запит
        cursor.execute(query)
        
        # Запускаємо індикатор завантаження в окремому потоці
        # Не виводимо дубльоване повідомлення, оскільки воно буде в анімованому індикаторі
        
        # Оцінка часу виконання запиту, використовуючи усереднене значення у 5 хвилин
        # Ви можете налаштувати це значення на основі ваших спостережень
        estimated_query_time = 120  # 5 хвилин у секундах
        
        # Якщо є глобальна змінна з інформацією про середній час запитів, використовуємо її
        global avg_query_time
        if 'avg_query_time' in globals() and avg_query_time is not None:
            estimated_query_time = avg_query_time
        
        spinner_thread = threading.Thread(
            target=loading_spinner, 
            args=("Отримання даних з OLAP кубу", estimated_query_time)
        )
        spinner_thread.daemon = True
        spinner_thread.start()
        
        try:
            # Отримуємо всі рядки відразу
            rows = cursor.fetchall()
            # Зупиняємо анімацію
            animation_running = False
            spinner_thread.join(timeout=1.0)
            
            # Отримуємо імена колонок
            columns = [desc[0] for desc in cursor.description]
            
            query_end_time = time.time()
            query_duration = query_end_time - query_start_time
            
            # Оновлюємо середній час виконання запиту
            if 'avg_query_time' not in globals() or avg_query_time is None:
                avg_query_time = query_duration
            else:
                # Плавне оновлення середнього часу (алгоритм експоненційного згладжування)
                # Alpha - коефіцієнт згладжування (0.3 означає, що новий вимір має вагу 30%)
                alpha = 0.3
                avg_query_time = (1 - alpha) * avg_query_time + alpha * query_duration
            
            print_success(f"Запит виконано за {format_time(query_duration)}. Отримано {len(rows)} рядків даних.")
            
            cursor.close()
            
            # Створюємо DataFrame з отриманих даних
            df = pd.DataFrame(rows, columns=columns)
            
            # Якщо немає даних, повертаємо порожній список
            if len(df) == 0:
                print_warning(f"Запит не повернув даних для періоду {reporting_period}")
                return []
            
            print_progress(f"Обробка результатів запиту...")
            # Перейменовуємо стовпці для відповідності формату DAX Studio
            renamed_columns = {}
            potential_names = {}
            
            # Перший прохід: збираємо потенційні імена і перевіряємо дублікати
            for col in df.columns:
                # Шаблон для розпізнавання стовпців у форматі "TableName[ColumnName]"
                match = re.match(r'(\w+)\[([^\]]+)\]', col)
                if match:
                    # Витягуємо тільки назву стовпця без таблиці та дужок
                    column_name = match.group(2)
                    if column_name in potential_names:
                        # Дублювання виявлено, позначаємо обидва стовпці для збереження оригінальних назв
                        potential_names[column_name] = False
                    else:
                        # Поки що унікальне ім'я, помічаємо як потенційно перейменоване
                        potential_names[column_name] = True
                else:
                    # Для обчислюваних стовпців просто видаляємо квадратні дужки
                    column_name = col.strip('[]')
                    # Їхні імена зазвичай унікальні, але все одно перевіряємо
                    if column_name in potential_names:
                        potential_names[column_name] = False
                    else:
                        potential_names[column_name] = True
            
            # Другий прохід: застосовуємо перейменування, уникаючи дублікатів
            for col in df.columns:
                match = re.match(r'(\w+)\[([^\]]+)\]', col)
                if match:
                    column_name = match.group(2)
                    # Перейменовуємо тільки якщо немає конфлікту імен
                    if potential_names[column_name]:
                        renamed_columns[col] = column_name
                    # Інакше залишаємо оригінальну назву
                else:
                    # Для обчислюваних стовпців завжди видаляємо квадратні дужки
                    renamed_columns[col] = col.strip('[]')
            
            # Виводимо інформацію про стовпці, які не були перейменовані через дублювання
            duplicate_columns = [col for col in df.columns if re.match(r'(\w+)\[([^\]]+)\]', col) and 
                            re.match(r'(\w+)\[([^\]]+)\]', col).group(2) in potential_names and 
                            not potential_names[re.match(r'(\w+)\[([^\]]+)\]', col).group(2)]]
            
            if duplicate_columns:
                print_warning(f"Деякі стовпці не були перейменовані через потенційне дублювання:")
                for col in duplicate_columns:
                    match = re.match(r'(\w+)\[([^\]]+)\]', col)
                    if match:
                        print(f"   {Fore.YELLOW}• {Fore.WHITE}{col} {Fore.YELLOW}(конфлікт імені: {Fore.WHITE}{match.group(2)}{Fore.YELLOW})")
            else:
                print_info("Усі стовпці успішно перейменовано")
            
            # Застосовуємо нові назви стовпців
            df.rename(columns=renamed_columns, inplace=True)
            
            # Експортуємо дані у Excel-файл з форматуванням
            print_progress(f"Експорт даних у Excel-файл {filepath}...")
            
            # Спочатку створюємо Excel-файл з даними
            df.to_excel(filepath, index=False)
            
            # Тепер відкриваємо його за допомогою openpyxl для форматування
            from openpyxl import load_workbook
            
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Налаштування стилів для заголовка з .env
            header_font = Font(
                name='Arial', 
                size=int(os.getenv('EXCEL_HEADER_FONT_SIZE', 11)), 
                bold=True, 
                color=os.getenv('EXCEL_HEADER_FONT_COLOR', 'FFFFFF')
            )
            header_fill = PatternFill(
                start_color=os.getenv('EXCEL_HEADER_COLOR', '00365E'), 
                end_color=os.getenv('EXCEL_HEADER_COLOR', '00365E'), 
                fill_type='solid'
            )
            
            # Застосування стилів до заголовків
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Закріплення заголовка, щоб він завжди був видимий при прокрутці
            ws.freeze_panes = 'A2'  # Закріплюємо перший рядок
            
            # Автоматичне налаштування ширини стовпців
            # Перебираємо всі стовпці та знаходимо максимальну довжину значення
            for col in range(1, len(df.columns) + 1):
                column_width = max(
                    len(str(df.columns[col-1])),  # Довжина заголовка
                    df.iloc[:, col-1].astype(str).str.len().max()  # Максимальна довжина даних
                )
                # Обмежуємо максимальну ширину стовпця
                adjusted_width = min(column_width + 2, 50)  # +2 для відступів
                ws.column_dimensions[get_column_letter(col)].width = adjusted_width
            
            # Зберігаємо відформатований файл
            wb.save(filepath)
            
            # Отримуємо розмір файлу та форматуємо його для виведення
            file_size_bytes = os.path.getsize(filepath)
            if file_size_bytes < 1024 * 1024:  # Менше 1 МБ
                file_size = f"{file_size_bytes / 1024:.1f} КБ"
            else:  # Більше або рівно 1 МБ
                file_size = f"{file_size_bytes / (1024 * 1024):.2f} МБ"
            
            print_success(f"Дані експортовано у файл: {Fore.WHITE}{filepath} {Fore.YELLOW}({file_size}, {len(df)} рядків)")
            
            # Повертаємо шлях до файлу для підтвердження успішного створення
            return filepath
            
        except Exception as e:
            # Зупиняємо анімацію при помилці
            animation_running = False
            spinner_thread.join(timeout=1.0)
            raise e
        
    except Exception as e:
        print_error(f"Помилка при виконанні запиту: {e}")
        return None

# Функція для отримання доступних тижнів з куба OLAP
def get_available_weeks(connection):
    """Отримує список доступних тижнів з куба OLAP"""
    print_info("Отримання доступних тижнів з куба OLAP...")
    
    query = """
    /* START QUERY BUILDER */
    EVALUATE
    SUMMARIZECOLUMNS(
        'Calendar'[year_num],
        'Calendar'[week_num]
    )
    ORDER BY 
        'Calendar'[year_num] ASC,
        'Calendar'[week_num] ASC
    /* END QUERY BUILDER */
    """
    
    try:
        cursor = connection.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        cursor.close()
        
        available_weeks = []
        for row in rows:
            year = int(row[0])  # year_num
            week = int(row[1])  # week_num
            available_weeks.append((year, week))
        
        print_info(f"Отримано {len(available_weeks)} доступних тижнів з куба")
        return available_weeks
    
    except Exception as e:
        print_error(f"Помилка при отриманні доступних тижнів: {e}")
        return []

# Функція для відображення зворотнього відліку
def countdown_timer(seconds):
    """Відображає зворотній відлік"""
    for remaining in range(seconds, 0, -1):
        # Форматуємо час, що залишився
        time_left = format_time(remaining)
        sys.stdout.write(f"\r{Fore.YELLOW}[{get_current_time()}] ⏱️ Очікування: залишилось {time_left}...")
        sys.stdout.flush()
        time.sleep(1)
    print()  # Переходимо на новий рядок після завершення

# Головний код
try:
    # Отримуємо параметри з .env файлу
    load_dotenv()
    
    print_header(f"OLAP ЕКСПОРТ ДАНИХ - НАЛАШТУВАННЯ")
    
    # Зчитуємо періоди з .env файлу
    start_period = os.getenv('YEAR_WEEK_START')
    end_period = os.getenv('YEAR_WEEK_END')
    
    # Ініціалізація підключення до OLAP
    connection_string, auth_details = get_connection_string()
    connection = connect_to_olap(connection_string, auth_details)
    if not connection:
        print_error("Не вдалося підключитися до OLAP. Програма завершує роботу.")
        sys.exit(1)

    # Отримуємо доступні тижні з куба
    available_weeks = get_available_weeks(connection)
    
    # Якщо періоди вказані, генеруємо список пар (рік, тиждень)
    if start_period and end_period:
        year_week_pairs = generate_year_week_pairs(start_period, end_period, available_weeks)
        if not year_week_pairs:
            print_error("Не вдалося згенерувати список періодів. Використовуються значення за замовчуванням.")
            year_num = CURRENT_YEAR
            week_nums = [CURRENT_WEEK]
            year_week_pairs = [(year_num, week) for week in week_nums]
    else:
        # Задаємо значення для року та списку тижнів за замовчуванням
        year_num = CURRENT_YEAR
        week_nums = [CURRENT_WEEK]  # Список тижнів для обробки
        year_week_pairs = [(year_num, week) for week in week_nums]
    
    filter_fg1_name = os.getenv('FILTER_FG1_NAME')

    # Створюємо структуру папок для збереження результатів
    result_dir = "result"

    # Перевіряємо і створюємо основну папку, якщо вона не існує
    if not os.path.exists(result_dir):
        os.makedirs(result_dir)
        print_info(f"Створено директорію '{result_dir}'")

    # Попередньо створюємо всі папки для років, які будуть використовуватись
    for year, _ in set((year, 0) for year, _ in year_week_pairs):
        year_dir = os.path.join(result_dir, str(year))
        if not os.path.exists(year_dir):
            os.makedirs(year_dir)
            print_info(f"Створено директорію '{year_dir}'")

    # Зчитуємо налаштування таймауту між запитами
    query_timeout = int(os.getenv('QUERY_TIMEOUT', 30))  # Значення за замовчуванням 30 секунд

    # Виводимо інформацію про параметри запуску
    print_header(f"OLAP ЕКСПОРТ ДАНИХ - ПОЧАТОК РОБОТИ")
    print_info(f"Налаштування:")
    print(f"   {Fore.CYAN}OLAP сервер:  {Fore.WHITE}{os.getenv('OLAP_SERVER')}")
    print(f"   {Fore.CYAN}База даних:   {Fore.WHITE}{os.getenv('OLAP_DATABASE')}")
    print(f"   {Fore.CYAN}Фільтр:       {Fore.WHITE}{filter_fg1_name}")
    
    # Додаємо інформацію про метод автентифікації
    auth_method = os.getenv('OLAP_AUTH_METHOD', AUTH_SSPI).upper()
    if auth_method == AUTH_SSPI:
        print(f"   {Fore.CYAN}Автентифікація: {Fore.WHITE}Windows (SSPI) як користувач {get_current_windows_user()}")
    elif auth_method == AUTH_LOGIN:
        user = os.getenv('OLAP_USER')
        if HAS_PYWIN32:
            print(f"   {Fore.CYAN}Автентифікація: {Fore.WHITE}Логін/пароль як користувач {user} через ADO (OLE DB)")
        else: 
            print(f"   {Fore.CYAN}Автентифікація: {Fore.WHITE}Логін/пароль як користувач {user} через ADOMD.NET (потрібен pywin32)")
    else:
        print(f"   {Fore.CYAN}Автентифікація: {Fore.WHITE}Невідомий метод ({auth_method})")
    
    # Виводимо інформацію про періоди
    if start_period and end_period:
        print(f"   {Fore.CYAN}Період:       {Fore.WHITE}з {start_period} по {end_period}")
        print(f"   {Fore.CYAN}Кількість періодів: {Fore.WHITE}{len(year_week_pairs)}")
    else:
        print(f"   {Fore.CYAN}Рік:          {Fore.WHITE}{year_num}")
        print(f"   {Fore.CYAN}Тижні:        {Fore.WHITE}{', '.join(map(str, week_nums))}")
    
    print(f"   {Fore.CYAN}Таймаут:      {Fore.WHITE}{query_timeout} секунд")
    
    # Початок відліку часу
    start_time = time.time()
    
    # Виконуємо запити для всіх тижнів
    files_created = []
    
    print_info(f"Запуск обробки для {len(year_week_pairs)} тижнів...")
    
    # Ініціалізуємо трекер часу
    time_tracker = TimeTracker(len(year_week_pairs))
    
    for i, (year, week) in enumerate(year_week_pairs):
        # Для першого тижня не робимо затримку
        if i > 0:
            print(f"\n{Fore.YELLOW}{'-' * 40}")
            print_info(f"Очікування {query_timeout} секунд перед наступним запитом...")
            
            # Починаємо відлік очікування
            time_tracker.start_waiting()
            
            # Виконуємо зворотний відлік
            countdown_timer(query_timeout)
            
            # Завершуємо відлік очікування
            time_tracker.end_waiting()
        
        reporting_period = f"{year}-{week:02d}"  # Формат РРРР-ТТ
        print(f"\n{Fore.CYAN}{'-' * 40}")
        
        # Відображаємо інформацію про прогрес обробки
        if i > 0:  # Після обробки хоча б одного елемента можемо показувати прогноз
            progress_info = time_tracker.get_progress_info()
            print(f"{Fore.MAGENTA}{progress_info}")
        
        print_info(f"Обробка тижня: {reporting_period} ({i+1}/{len(year_week_pairs)})")
        
        # Виконуємо запит і отримуємо результати
        file_path = run_mdx_query(connection, reporting_period)
        
        # Додаємо шлях до файлу до списку створених файлів
        if file_path:
            files_created.append(file_path)
        
        # Оновлюємо трекер часу після обробки елемента
        time_tracker.update()
    
    # Завершення відліку часу
    end_time = time.time()
    processing_time = end_time - start_time
    
    # Виводимо підсумок обробки
    print_header(f"ПІДСУМОК ОБРОБКИ")
    # Детальна інформація про час виконання
    if len(year_week_pairs) > 1:
        avg_time_per_week = processing_time / len(year_week_pairs)
        print_info(f"Деталі часу виконання:")
        print(f"   {Fore.CYAN}Загальний час:    {Fore.WHITE}{format_time(processing_time)}")
        print(f"   {Fore.CYAN}Середній час на 1 тиждень: {Fore.WHITE}{format_time(avg_time_per_week)}")
        if time_tracker.elapsed_times:
            min_time = min(time_tracker.elapsed_times)
            max_time = max(time_tracker.elapsed_times)
            print(f"   {Fore.CYAN}Мінімальний час: {Fore.WHITE}{format_time(min_time)}")
            print(f"   {Fore.CYAN}Максимальний час: {Fore.WHITE}{format_time(max_time)}")
    else:
        print_success(f"Обробку завершено за {format_time(processing_time)}")

    print_info(f"Створено файлів: {len(files_created)}")
    
    if files_created:
        for i, file_path in enumerate(files_created, 1):
            file_size_bytes = os.path.getsize(file_path)
            if file_size_bytes < 1024 * 1024:  # Менше 1 МБ
                file_size = f"{file_size_bytes / 1024:.1f} КБ"
            else:  # Більше або рівно 1 МБ
                file_size = f"{file_size_bytes / (1024 * 1024):.2f} МБ"
            print(f"   {Fore.CYAN}{i}. {Fore.WHITE}{file_path} {Fore.YELLOW}({file_size})")
    else:
        print_warning("Не було створено жодного файлу")
    
    # Закриваємо підключення до OLAP
    if connection:
        connection.close()
        print_info("Підключення до OLAP сервера закрито")

except Exception as e:
    print_error(f"Помилка при виконанні програми: {e}")
    sys.exit(1)

finally:
    # Переконуємось, що анімація зупинена
    animation_running = False